import openpyxl
from operator import itemgetter

def avgGrd(List):
    counter = 0
    num = List[0]
    for i in List:
        frq = List.count(i)
        if (frq > counter):
            counter = frq
            num = i
    return num

def getData(path):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    allSub = []
    Subj = []
    max_row = sheet_obj.max_row
    #taking all subject grades in allSub variable as list of list
    for j in range(7, 14):
        Subj = []
        for i in range(2, max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            Subj.append(cell_obj.value)
        allSub.append(Subj)

    #replacing "F (ABS)" to 'F'
    for i, x in enumerate(allSub):
        for j, a in enumerate(x):
            if 'F (ABS)' in a:
                allSub[i][j] = a.replace('F (ABS)', 'F')

    print("gotdata")
    data = analyse(allSub)
    appnd(wb_obj,data,path,allSub)


def analyse(allSub):
    grads = ["A+", "A", "B+", "B", "C+", "C", "D", "F", "I", "W"]
    analysis = []
    rslt=[]
    

    #counting grades 
    print(len(allSub))
    for i in allSub:
        SUBJ = []       
        for grd in grads:
            SUBJ.append(i.count(grd))
        rslt.append(SUBJ)
    
    #making new list of list with indexed values of list...list of 1st's 1st and 2nd's 2nd...    
    for i in range(10):
        analysis.append(list(map(itemgetter(i), rslt)))
    return analysis 



def appnd(wb_obj,data,path,allsub):
    sheet_obj = wb_obj.active
    grads = ["A+", "A", "B+", "B", "C+", "C", "D", "F", "I", "W"]
    headers = ["GRADE"]
    max_column = sheet_obj.max_column   
    #retrieving headers from excel
    for i in range(7, max_column + 1):
        cell_obj = sheet_obj.cell(row=1, column=i)
        headers.append(cell_obj.value)
    wb_obj.save(path)
  
    #inserting headers 
    r=2
    colmn=16
    for row in headers:
        sheet_obj.cell(row=r, column=colmn).value = row
        colmn+=1

    #inserting grades 
    r=3
    colmn=16    
    for row in grads:
        sheet_obj.cell(row=r, column=colmn).value = row
        r+=1
        
    #inserting data  
    r = 3
    for innerlist in data: 
        colmn=17
        for row in innerlist:
            sheet_obj.cell(row=r, column=colmn).value = row
            colmn+=1
        r+= 1
    #avg grade
    avgrd = ["Avg"]
    for list in allsub:
        avgrd.append(avgGrd(list))
      
    #inserting grades
    r=13
    colmn=16
    for row in avgrd:
        sheet_obj.cell(row=r, column=colmn).value = row
        colmn+=1
    wb_obj.save(path)
    print("data saved")




    
